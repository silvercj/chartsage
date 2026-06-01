import io, zipfile
from schemas import Report, ChartWithCaption, ChartSpec, ChartLayoutEntry, KeyMetric
import report_export as rx

def _report():
    spec = ChartSpec(kind="bar", title="By region", intent="i", x=["W","E"], y=[2,1], source_columns=["region"], data_point_count=2)
    return Report(generated_at="2024-01-01T00:00:00", summary="Sales summary.", data_quality=["1% blanks"],
                  charts=[ChartWithCaption(chart_id="c1", spec=spec, caption="West leads.")],
                  layout=[ChartLayoutEntry(chart_id="c1", position="main", order=0)],
                  key_metrics=[KeyMetric(label="Total", value=1234.0, format="currency")])

IMAGES = [{"chart_id": "c1", "png": b"\x89PNG\r\n\x1a\n" + b"0"*64}]  # not a real PNG; builders just embed bytes

def test_pptx_opens_with_slides():
    from pptx import Presentation
    data = rx.build_pptx(_report(), IMAGES)
    prs = Presentation(io.BytesIO(data))
    assert len(prs.slides) >= 1 + 1 + len(IMAGES)   # title + summary/kpi + one per chart (exact count per impl)

def test_xlsx_has_data_and_summary_sheets():
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(rx.build_xlsx(_report(), b"region,sales\nW,2\nE,1\n")))
    assert set(wb.sheetnames) >= {"Data", "Summary"}

def test_png_zip_has_one_entry_per_chart():
    z = zipfile.ZipFile(io.BytesIO(rx.build_png_zip(IMAGES)))
    assert len(z.namelist()) == len(IMAGES)

def test_markdown_contains_summary_and_kpis():
    md = rx.build_markdown(_report(), IMAGES)
    assert "Sales summary." in md and "Total" in md and "data:image/png;base64," in md
